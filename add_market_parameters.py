"""
Script per aggiungere i nuovi parametri TAM/SAM/SOM all'Excel.

Aggiunge nella sezione Assumptions:
- Market_Max_Followers_Local (50000)
- Market_Max_Followers_Global (1000000)
- Market_Max_PayingUsers_Local (2000)
- Market_Max_PayingUsers_Global (25000)
- Follower_Adoption_Ramp_Months (24)
"""

import openpyxl

def add_market_parameters():
    excel_path = r'c:\Users\simia\Desktop\Business_analysis\ai_finance_dynamic_model_v7_channels.xlsx'
    
    print(f"Opening Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Model']
    
    # Trova l'ultima riga delle assumptions (prima della sezione vuota)
    last_assumption_row = None
    for row_idx in range(3, 100):
        cell_b = ws.cell(row=row_idx, column=2)  # Column B = Parameter
        if cell_b.value is None or str(cell_b.value).strip() == '':
            last_assumption_row = row_idx - 1
            break
    
    if last_assumption_row is None:
        print("ERROR: Could not find end of assumptions section")
        return
    
    print(f"Last assumption row: {last_assumption_row}")
    
    # Controlla se i parametri esistono già
    existing_params = set()
    for row_idx in range(3, last_assumption_row + 1):
        param = ws.cell(row=row_idx, column=2).value
        if param:
            existing_params.add(param)
    
    # Parametri da aggiungere
    new_params = [
        {
            'category': 'MarketCaps',
            'parameter': 'Market_Max_Followers_Local',
            'value': 50000,
            'unit': 'followers',
            'notes': 'Max follower raggiungibili nel mercato nicchia Zurigo/Svizzera'
        },
        {
            'category': 'MarketCaps',
            'parameter': 'Market_Max_Followers_Global',
            'value': 1000000,
            'unit': 'followers',
            'notes': 'Max follower raggiungibili a livello internazionale (espansione)'
        },
        {
            'category': 'MarketCaps',
            'parameter': 'Market_Max_PayingUsers_Local',
            'value': 2000,
            'unit': 'users',
            'notes': 'Max paying users nel mercato nicchia Zurigo/Svizzera (~10-20% hardcore)'
        },
        {
            'category': 'MarketCaps',
            'parameter': 'Market_Max_PayingUsers_Global',
            'value': 25000,
            'unit': 'users',
            'notes': 'Max paying users a livello internazionale'
        },
        {
            'category': 'MarketCaps',
            'parameter': 'Follower_Adoption_Ramp_Months',
            'value': 24,
            'unit': 'months',
            'notes': 'Mesi necessari per raggiungere il massimo potenziale di crescita (brand nuovo)'
        }
    ]
    
    # Filtra parametri già esistenti
    params_to_add = [p for p in new_params if p['parameter'] not in existing_params]
    
    if not params_to_add:
        print("\nAll parameters already exist in Excel. Nothing to add.")
        wb.close()
        return
    
    print(f"\nAdding {len(params_to_add)} new parameters:")
    
    # Inserisci nuovi parametri dopo l'ultimo esistente
    insert_row = last_assumption_row + 1
    
    for param in params_to_add:
        print(f"  Row {insert_row}: {param['parameter']} = {param['value']}")
        
        ws.cell(row=insert_row, column=1, value=param['category'])  # A: Category
        ws.cell(row=insert_row, column=2, value=param['parameter'])  # B: Parameter
        ws.cell(row=insert_row, column=3, value=param['value'])      # C: Value
        ws.cell(row=insert_row, column=4, value=param['unit'])       # D: Unit
        ws.cell(row=insert_row, column=5, value=param['notes'])      # E: Notes
        
        insert_row += 1
    
    # Salva
    print(f"\nSaving Excel file...")
    wb.save(excel_path)
    wb.close()
    
    print(f"\n✓ Successfully added {len(params_to_add)} new TAM/SAM/SOM parameters!")
    print(f"  Total assumptions now: {last_assumption_row - 2 + len(params_to_add)}")


if __name__ == "__main__":
    add_market_parameters()
