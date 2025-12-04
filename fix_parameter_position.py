"""
Script per spostare FollowerAds_CTR_to_Site dalla riga 98 alla riga corretta (dopo 48).
"""

import openpyxl
from openpyxl.styles import Alignment

def fix_parameter_position():
    """Sposta FollowerAds_CTR_to_Site alla posizione corretta."""
    
    filepath = 'ai_finance_dynamic_model_v7_channels.xlsx'
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    print("📂 Fixing parameter position in Excel...")
    print(f"   File: {filepath}")
    
    # Leggi il parametro dalla riga 98
    category_98 = ws.cell(98, 1).value
    param_98 = ws.cell(98, 2).value
    value_98 = ws.cell(98, 3).value
    unit_98 = ws.cell(98, 4).value
    notes_98 = ws.cell(98, 5).value
    
    print(f"\n📍 Trovato alla riga 98:")
    print(f"   {param_98} = {value_98}")
    
    # Inserisci alla riga 49 (dopo Follower_Threshold_For_Click_Ads che è alla 48)
    target_row = 49
    
    print(f"\n➡️  Spostamento alla riga {target_row}...")
    
    # Scrivi alla riga target
    ws.cell(target_row, 1).value = 'Ads'  # Category
    ws.cell(target_row, 2).value = param_98
    ws.cell(target_row, 3).value = value_98
    ws.cell(target_row, 4).value = unit_98
    ws.cell(target_row, 5).value = 'CTR from Follower Ads campaigns to website (1%)'
    
    # Allineamento
    for col in range(1, 6):
        ws.cell(target_row, col).alignment = Alignment(
            horizontal='left',
            vertical='center'
        )
    
    # Cancella dalla riga 98
    for col in range(1, 6):
        ws.cell(98, col).value = None
    
    print(f"✓ Parametro spostato alla riga {target_row}")
    
    # Salva
    output_path = filepath.replace('.xlsx', '_FIXED.xlsx')
    wb.save(output_path)
    
    print(f"\n✅ Excel salvato come: {output_path}")
    print("⚠️  Rinomina il file da _FIXED.xlsx al nome originale")
    
    return True

if __name__ == '__main__':
    fix_parameter_position()
