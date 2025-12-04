#!/usr/bin/env python3
"""
Script per aggiungere i nuovi parametri Paid Ads nell'Excel v7
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

excel_path = 'ai_finance_dynamic_model_v7_channels.xlsx'

print("=" * 80)
print("AGGIORNAMENTO EXCEL v7 - NUOVI PARAMETRI PAID ADS")
print("=" * 80)

# Carica workbook
print(f"\nCaricamento {excel_path}...")
wb = load_workbook(excel_path)
sheet = wb['Model']

# Trova l'ultima riga con assumptions (dovrebbe essere riga 46)
last_assumption_row = 46

# Nuovi parametri da aggiungere
new_params = [
    {
        'Category': 'Paid Social Ads',
        'Parameter': 'ClickAds_CPC_EUR',
        'Value': 2.0,
        'Unit': 'EUR per click',
        'Notes': 'Costo medio per click per campagne link-click (Fase 2)'
    },
    {
        'Category': 'Paid Social Ads',
        'Parameter': 'Follower_Threshold_For_Click_Ads',
        'Value': 20000,
        'Unit': 'followers',
        'Notes': 'Soglia followers per switch da Follower Ads (Fase 1) a Click Ads (Fase 2)'
    }
]

# Controlla se i parametri esistono già
print("\nVerifica parametri esistenti...")
existing_params = []
for row in range(4, 50):  # Scansiona righe assumptions
    param_name = sheet.cell(row, 2).value
    if param_name:
        existing_params.append(str(param_name))

params_to_add = []
for param in new_params:
    if param['Parameter'] not in existing_params:
        params_to_add.append(param)
        print(f"  ✓ {param['Parameter']} - DA AGGIUNGERE")
    else:
        print(f"  → {param['Parameter']} - GIÀ PRESENTE")

if not params_to_add:
    print("\n✓ Tutti i parametri sono già presenti nell'Excel!")
    print("Nessuna modifica necessaria.")
else:
    print(f"\nAggiunta di {len(params_to_add)} nuovi parametri...")
    
    # Aggiungi i nuovi parametri dopo l'ultima assumption
    current_row = last_assumption_row + 1
    
    for param in params_to_add:
        sheet.cell(current_row, 1).value = param['Category']
        sheet.cell(current_row, 2).value = param['Parameter']
        sheet.cell(current_row, 3).value = param['Value']
        sheet.cell(current_row, 4).value = param['Unit']
        sheet.cell(current_row, 5).value = param['Notes']
        
        print(f"  ✓ Riga {current_row}: {param['Parameter']} = {param['Value']}")
        current_row += 1
    
    # Salva il file
    output_path = 'ai_finance_dynamic_model_v7_channels_updated.xlsx'
    print(f"\nSalvataggio in {output_path}...")
    wb.save(output_path)
    
    print(f"\n✅ Excel aggiornato con successo!")
    print(f"\nFile creato: {output_path}")
    print("\nProssimi passi:")
    print("  1. Rinomina il file in 'ai_finance_dynamic_model_v7_channels.xlsx'")
    print("  2. Riavvia l'app: python financial_model_app_v2.py")
    print("  3. Clicca 'Recalculate' per applicare i nuovi parametri")

print("\n" + "=" * 80)
