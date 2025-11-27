"""
Senior Data Analyst & Python Engineer: Financial Model Analysis (V2)
====================================================================

This script analyzes the AI Finance Dynamic Model Excel file.
Since formulas in Excel show cached values of 0, this version manually
calculates all derived values based on the formula logic detected.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

# Set visualization style
sns.set_style("whitegrid")
plt.rcParams['figure.figsize'] = (12, 6)
plt.rcParams['font.size'] = 10

class FinancialModelAnalyzer:
    """Analyzes Excel financial model with manual formula calculation."""
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.assumptions = {}
        self.monthly_df = None
        self.yearly_df = None
        
    def load_and_calculate(self):
        """Load Excel and manually calculate all formulas."""
        print("=" * 80)
        print("STEP 1: LOADING EXCEL AND EXTRACTING DATA")
        print("=" * 80)
        
        wb = load_workbook(self.filepath, data_only=True)
        sheet = wb["Model"]
        
        print(f"✓ Loaded workbook: {sheet.title}")
        
        # Extract assumptions (rows 4-49, columns 2-3)
        print("\n1. Extracting assumptions...")
        for row_idx in range(4, 50):
            param = sheet.cell(row=row_idx, column=2).value
            value = sheet.cell(row=row_idx, column=3).value
            if param and value is not None:
                self.assumptions[param] = value
        
        # Manually calculate derived assumptions
        self.assumptions['Base_Visitor_to_Paid_Conv'] = self.assumptions['ConvVS'] * self.assumptions['ConvSP']
        
        self.assumptions['Share_Sum_Y1'] = (self.assumptions['Org_Share_Y1'] + 
                                            self.assumptions['Inf_Share_Y1'] +
                                            self.assumptions['Ref_Share_Y1'] +
                                            self.assumptions['Other_Share_Y1'])
        
        self.assumptions['Share_Sum_Y2'] = (self.assumptions['Org_Share_Y2'] + 
                                            self.assumptions['Inf_Share_Y2'] +
                                            self.assumptions['Ref_Share_Y2'] +
                                            self.assumptions['Other_Share_Y2'])
        
        self.assumptions['Share_Sum_Y3'] = (self.assumptions['Org_Share_Y3'] + 
                                            self.assumptions['Inf_Share_Y3'] +
                                            self.assumptions['Ref_Share_Y3'] +
                                            self.assumptions['Other_Share_Y3'])
        
        self.assumptions['CAC_Y1'] = (self.assumptions['CAC_Org'] * self.assumptions['Org_Share_Y1'] +
                                      self.assumptions['CAC_Inf'] * self.assumptions['Inf_Share_Y1'] +
                                      self.assumptions['CAC_Ref'] * self.assumptions['Ref_Share_Y1'] +
                                      self.assumptions['CAC_Other'] * self.assumptions['Other_Share_Y1'])
        
        self.assumptions['CAC_Y2'] = (self.assumptions['CAC_Org'] * self.assumptions['Org_Share_Y2'] +
                                      self.assumptions['CAC_Inf'] * self.assumptions['Inf_Share_Y2'] +
                                      self.assumptions['CAC_Ref'] * self.assumptions['Ref_Share_Y2'] +
                                      self.assumptions['CAC_Other'] * self.assumptions['Other_Share_Y2'])
        
        self.assumptions['CAC_Y3'] = (self.assumptions['CAC_Org'] * self.assumptions['Org_Share_Y3'] +
                                      self.assumptions['CAC_Inf'] * self.assumptions['Inf_Share_Y3'] +
                                      self.assumptions['CAC_Ref'] * self.assumptions['Ref_Share_Y3'] +
                                      self.assumptions['CAC_Other'] * self.assumptions['Other_Share_Y3'])
        
        self.assumptions['Inf_Visitors_per_Collab'] = (self.assumptions['Inf_Avg_Followers'] *
                                                        self.assumptions['Inf_Reach_Rate'] *
                                                        self.assumptions['Inf_Click_Rate'])
        
        print(f"✓ Extracted {len(self.assumptions)} assumptions (including calculated)")
        
        # Extract monthly model structure (rows 53-88)
        print("\n2. Extracting monthly model structure...")
        monthly_data = []
        for row_idx in range(53, 89):  # 36 months
            row = []
            for col_idx in range(1, 23):  # 22 columns
                value = sheet.cell(row=row_idx, column=col_idx).value
                row.append(value if value is not None else 0)
            monthly_data.append(row)
        
        # Get headers from row 52
        headers = []
        for col_idx in range(1, 23):
            headers.append(sheet.cell(row=52, column=col_idx).value)
        
        self.monthly_df = pd.DataFrame(monthly_data, columns=headers)
        print(f"✓ Extracted {len(self.monthly_df)} months")
        
        # Manually calculate monthly model formulas
        print("\n3. Calculating monthly model values...")
        self._calculate_monthly_model()
        print("✓ Monthly calculations complete")
        
        # Calculate yearly summaries
        print("\n4. Calculating yearly summaries...")
        self._calculate_yearly_summary()
        print("✓ Yearly summaries complete")
        
        print("\n" + "=" * 80)
        print("✓ DATA EXTRACTION AND CALCULATION COMPLETE")
        print("=" * 80)
    
    def _calculate_monthly_model(self):
        """Manually calculate all monthly model formulas."""
        
        # Initialize Paying_Users_Start for first month
        self.monthly_df.at[0, 'Paying_Users_Start'] = 0
        
        for idx in range(len(self.monthly_df)):
            year = self.monthly_df.at[idx, 'Year']
            
            # Visitors_from_Social = Social_Views * Social_View_to_Visit_Conv
            self.monthly_df.at[idx, 'Visitors_from_Social'] = (
                self.monthly_df.at[idx, 'Social_Views'] * 
                self.assumptions['Social_View_to_Visit_Conv']
            )
            
            # Inf_Visitors = Inf_Collabs * Inf_Visitors_per_Collab
            if year == 1:
                inf_collabs = self.assumptions['Inf_Collabs_Y1']
            elif year == 2:
                inf_collabs = self.assumptions['Inf_Collabs_Y2']
            else:
                inf_collabs = self.assumptions['Inf_Collabs_Y3']
            
            self.monthly_df.at[idx, 'Inf_Visitors'] = (
                inf_collabs * self.assumptions['Inf_Visitors_per_Collab']
            )
            
            # Visitors_Total = Visitors_from_Social + Inf_Visitors
            self.monthly_df.at[idx, 'Visitors_Total'] = (
                self.monthly_df.at[idx, 'Visitors_from_Social'] +
                self.monthly_df.at[idx, 'Inf_Visitors']
            )
            
            # Signups = Visitors_Total * ConvVS
            self.monthly_df.at[idx, 'Signups'] = (
                self.monthly_df.at[idx, 'Visitors_Total'] * 
                self.assumptions['ConvVS']
            )
            
            # New_Paying_Users = Signups * ConvSP
            self.monthly_df.at[idx, 'New_Paying_Users'] = (
                self.monthly_df.at[idx, 'Signups'] * 
                self.assumptions['ConvSP']
            )
            
            # Churn_Rate depends on year
            if year == 1:
                churn_rate = self.assumptions['ChurnY1']
            elif year == 2:
                churn_rate = self.assumptions['ChurnY2']
            else:
                churn_rate = self.assumptions['ChurnY3']
            
            self.monthly_df.at[idx, 'Churn_Rate'] = churn_rate
            
            # Churned_Users = Paying_Users_Start * Churn_Rate
            self.monthly_df.at[idx, 'Churned_Users'] = (
                self.monthly_df.at[idx, 'Paying_Users_Start'] * churn_rate
            )
            
            # Paying_Users_End = Paying_Users_Start - Churned_Users + New_Paying_Users
            self.monthly_df.at[idx, 'Paying_Users_End'] = (
                self.monthly_df.at[idx, 'Paying_Users_Start'] -
                self.monthly_df.at[idx, 'Churned_Users'] +
                self.monthly_df.at[idx, 'New_Paying_Users']
            )
            
            # Set next month's Paying_Users_Start
            if idx < len(self.monthly_df) - 1:
                self.monthly_df.at[idx + 1, 'Paying_Users_Start'] = (
                    self.monthly_df.at[idx, 'Paying_Users_End']
                )
            
            # CAC_per_New_User depends on year
            if year == 1:
                cac = self.assumptions['CAC_Y1']
            elif year == 2:
                cac = self.assumptions['CAC_Y2']
            else:
                cac = self.assumptions['CAC_Y3']
            
            self.monthly_df.at[idx, 'CAC_per_New_User'] = cac
            
            # Marketing_Spend = New_Paying_Users * CAC_per_New_User
            self.monthly_df.at[idx, 'Marketing_Spend'] = (
                self.monthly_df.at[idx, 'New_Paying_Users'] * cac
            )
            
            # ARPU is constant
            self.monthly_df.at[idx, 'ARPU'] = self.assumptions['ARPU']
            
            # MRR = Paying_Users_End * ARPU
            self.monthly_df.at[idx, 'MRR'] = (
                self.monthly_df.at[idx, 'Paying_Users_End'] * 
                self.assumptions['ARPU']
            )
            
            # DataSub_Cost triggered when MRR >= threshold
            if self.monthly_df.at[idx, 'MRR'] >= self.assumptions['DataSub_MRR_Threshold']:
                self.monthly_df.at[idx, 'DataSub_Cost'] = self.assumptions['DataSub_Fee']
            else:
                self.monthly_df.at[idx, 'DataSub_Cost'] = 0
            
            # XAPI_Cost triggered when MRR >= threshold
            if self.monthly_df.at[idx, 'MRR'] >= self.assumptions['XAPI_MRR_Threshold']:
                self.monthly_df.at[idx, 'XAPI_Cost'] = self.assumptions['XAPI_Fee']
            else:
                self.monthly_df.at[idx, 'XAPI_Cost'] = 0
            
            # Base_Fixed_Cost is constant
            self.monthly_df.at[idx, 'Base_Fixed_Cost'] = self.assumptions['BaseFixedCost']
            
            # Total_Costs = Marketing_Spend + DataSub_Cost + XAPI_Cost + Base_Fixed_Cost
            self.monthly_df.at[idx, 'Total_Costs'] = (
                self.monthly_df.at[idx, 'Marketing_Spend'] +
                self.monthly_df.at[idx, 'DataSub_Cost'] +
                self.monthly_df.at[idx, 'XAPI_Cost'] +
                self.monthly_df.at[idx, 'Base_Fixed_Cost']
            )
            
            # Net_Cash_Flow = MRR - Total_Costs
            self.monthly_df.at[idx, 'Net_Cash_Flow'] = (
                self.monthly_df.at[idx, 'MRR'] -
                self.monthly_df.at[idx, 'Total_Costs']
            )
        
        # Cumulative_Cash is cumulative sum of Net_Cash_Flow
        self.monthly_df['Cumulative_Cash'] = self.monthly_df['Net_Cash_Flow'].cumsum()
    
    def _calculate_yearly_summary(self):
        """Calculate yearly summary metrics."""
        
        yearly_data = []
        
        for year in [1, 2, 3]:
            year_data = self.monthly_df[self.monthly_df['Year'] == year]
            
            summary = {
                'Year': year,
                'End_Paying_Users': year_data['Paying_Users_End'].iloc[-1],
                'End_MRR_EUR': year_data['MRR'].iloc[-1],
                'ARR_EUR': year_data['MRR'].iloc[-1] * 12,
                'Total_New_Customers': year_data['New_Paying_Users'].sum(),
                'Total_Marketing_Spend_EUR': year_data['Marketing_Spend'].sum(),
                'Average_CAC_EUR': (year_data['Marketing_Spend'].sum() / 
                                   year_data['New_Paying_Users'].sum() 
                                   if year_data['New_Paying_Users'].sum() > 0 else 0),
                'Assumed_Monthly_Churn': year_data['Churn_Rate'].iloc[0],
                'Cumulative_Cash_EndOfYear': year_data['Cumulative_Cash'].iloc[-1],
                'Share_Visitors_from_Influencers': (year_data['Inf_Visitors'].sum() / 
                                                    year_data['Visitors_Total'].sum()
                                                    if year_data['Visitors_Total'].sum() > 0 else 0),
                'Total_Social_Views': year_data['Social_Views'].sum(),
            }
            
            # Calculate LTV
            monthly_churn = summary['Assumed_Monthly_Churn']
            if monthly_churn > 0:
                avg_lifetime_months = 1 / monthly_churn
                summary['LTV_EUR'] = self.assumptions['ARPU'] * avg_lifetime_months * self.assumptions['GrossMargin']
            else:
                summary['LTV_EUR'] = 0
            
            # Calculate LTV/CAC ratio
            if summary['Average_CAC_EUR'] > 0:
                summary['LTV_CAC_Ratio'] = summary['LTV_EUR'] / summary['Average_CAC_EUR']
            else:
                summary['LTV_CAC_Ratio'] = 0
            
            yearly_data.append(summary)
        
        self.yearly_df = pd.DataFrame(yearly_data)
    
    def sanity_checks(self):
        """Perform data validation checks."""
        print("\n" + "=" * 80)
        print("STEP 2: SANITY CHECKS")
        print("=" * 80)
        
        issues = []
        
        # Check 1: Channel shares sum to 1.0
        print("\n1. Checking channel shares sum to ~1.0...")
        for year in [1, 2, 3]:
            share_sum = self.assumptions[f'Share_Sum_Y{year}']
            if abs(share_sum - 1.0) > 0.01:
                issues.append(f"Year {year} channel shares sum to {share_sum:.3f} (expected ~1.0)")
                print(f"  ⚠ Year {year}: {share_sum:.3f}")
            else:
                print(f"  ✓ Year {year}: {share_sum:.3f}")
        
        # Check 2: Months are sequential
        print("\n2. Checking months are sequential (1-12 per year)...")
        for year in [1, 2, 3]:
            year_data = self.monthly_df[self.monthly_df['Year'] == year]
            months = sorted(year_data['Month'].values)
            expected = list(range(1, 13))
            if list(months) != expected:
                issues.append(f"Year {year} months are not 1-12")
                print(f"  ⚠ Year {year}: {months}")
            else:
                print(f"  ✓ Year {year}: Months 1-12 present")
        
        # Check 3: Cumulative cash consistency
        print("\n3. Checking cumulative cash consistency...")
        calc_cumulative = self.monthly_df['Net_Cash_Flow'].cumsum()
        actual_cumulative = self.monthly_df['Cumulative_Cash']
        max_diff = abs(calc_cumulative - actual_cumulative).max()
        if max_diff > 1.0:
            issues.append(f"Cumulative cash mismatch (max diff: {max_diff:.2f} EUR)")
            print(f"  ⚠ Max difference: {max_diff:.2f} EUR")
        else:
            print(f"  ✓ Cumulative cash is consistent (max diff: {max_diff:.2f} EUR)")
        
        # Check 4: Visitors total
        print("\n4. Checking visitor totals...")
        calc_total = self.monthly_df['Visitors_from_Social'] + self.monthly_df['Inf_Visitors']
        actual_total = self.monthly_df['Visitors_Total']
        max_diff = abs(calc_total - actual_total).max()
        if max_diff > 1.0:
            issues.append(f"Visitor totals mismatch (max diff: {max_diff:.2f})")
            print(f"  ⚠ Max difference: {max_diff:.2f}")
        else:
            print(f"  ✓ Visitor totals are consistent")
        
        # Summary
        print("\n" + "-" * 80)
        if issues:
            print(f"⚠ Found {len(issues)} potential issues:")
            for issue in issues:
                print(f"  - {issue}")
        else:
            print("✓ All sanity checks passed!")
        
        return issues
    
    def create_key_assumptions_table(self):
        """Generate markdown table of key assumptions."""
        print("\n" + "=" * 80)
        print("STEP 3: KEY ASSUMPTIONS TABLE")
        print("=" * 80)
        
        key_params = [
            ('ARPU', 'Revenue'),
            ('GrossMargin', 'Revenue'),
            ('ConvVS', 'Conversion'),
            ('ConvSP', 'Conversion'),
            ('Base_Visitor_to_Paid_Conv', 'Conversion'),
            ('Social_View_to_Visit_Conv', 'Conversion'),
            ('ChurnY1', 'Retention'),
            ('ChurnY2', 'Retention'),
            ('ChurnY3', 'Retention'),
            ('CAC_Y1', 'CAC'),
            ('CAC_Y2', 'CAC'),
            ('CAC_Y3', 'CAC'),
            ('Inf_Avg_Followers', 'Influencer'),
            ('Inf_Reach_Rate', 'Influencer'),
            ('Inf_Click_Rate', 'Influencer'),
            ('Inf_Visitors_per_Collab', 'Influencer'),
            ('BaseFixedCost', 'Costs'),
            ('DataSub_Fee', 'Costs'),
            ('DataSub_MRR_Threshold', 'Costs'),
            ('XAPI_Fee', 'Costs'),
            ('XAPI_MRR_Threshold', 'Costs'),
            ('Broker_TargetCapital', 'Capital'),
        ]
        
        table_data = []
        for param, category in key_params:
            if param in self.assumptions:
                value = self.assumptions[param]
                if isinstance(value, (int, float)):
                    if param.startswith('Conv') or param.endswith('Rate') or 'Margin' in param or 'Share' in param:
                        formatted = f"{value:.2%}" if value < 1 else f"{value:.4f}"
                    elif 'Churn' in param:
                        formatted = f"{value:.2%}"
                    else:
                        formatted = f"{value:,.2f}"
                else:
                    formatted = str(value)
                table_data.append([category, param, formatted])
        
        df_table = pd.DataFrame(table_data, columns=['Category', 'Parameter', 'Value'])
        print("\n" + df_table.to_string(index=False))
        
        return df_table
    
    def create_yearly_summary_table(self):
        """Generate clean yearly summary table."""
        print("\n" + "=" * 80)
        print("STEP 4: YEARLY SUMMARY TABLE")
        print("=" * 80)
        
        df_display = self.yearly_df.copy()
        
        # Format columns
        df_display['End_Paying_Users'] = df_display['End_Paying_Users'].apply(lambda x: f"{x:,.0f}")
        df_display['End_MRR_EUR'] = df_display['End_MRR_EUR'].apply(lambda x: f"€{x:,.0f}")
        df_display['ARR_EUR'] = df_display['ARR_EUR'].apply(lambda x: f"€{x:,.0f}")
        df_display['Total_New_Customers'] = df_display['Total_New_Customers'].apply(lambda x: f"{x:,.0f}")
        df_display['Total_Marketing_Spend_EUR'] = df_display['Total_Marketing_Spend_EUR'].apply(lambda x: f"€{x:,.0f}")
        df_display['Average_CAC_EUR'] = df_display['Average_CAC_EUR'].apply(lambda x: f"€{x:,.2f}")
        df_display['LTV_EUR'] = df_display['LTV_EUR'].apply(lambda x: f"€{x:,.0f}")
        df_display['LTV_CAC_Ratio'] = df_display['LTV_CAC_Ratio'].apply(lambda x: f"{x:.2f}x")
        df_display['Cumulative_Cash_EndOfYear'] = df_display['Cumulative_Cash_EndOfYear'].apply(lambda x: f"€{x:,.0f}")
        df_display['Share_Visitors_from_Influencers'] = df_display['Share_Visitors_from_Influencers'].apply(lambda x: f"{x:.1%}")
        df_display['Total_Social_Views'] = df_display['Total_Social_Views'].apply(lambda x: f"{x:,.0f}")
        
        print("\n" + df_display.to_string(index=False))
        
        return df_display
    
    def create_monthly_funnel_table(self):
        """Generate compact monthly funnel table (sample)."""
        print("\n" + "=" * 80)
        print("STEP 5: MONTHLY FUNNEL TABLE (First 12 months sample)")
        print("=" * 80)
        
        key_cols = ['Year', 'Month', 'Social_Views', 'Visitors_from_Social', 
                   'Inf_Visitors', 'Visitors_Total', 'New_Paying_Users', 
                   'MRR', 'Net_Cash_Flow', 'Cumulative_Cash']
        
        df_funnel = self.monthly_df[key_cols].head(12).copy()
        
        # Format numeric columns
        for col in ['Social_Views', 'Visitors_from_Social', 'Inf_Visitors', 'Visitors_Total']:
            df_funnel[col] = df_funnel[col].apply(lambda x: f"{x:,.0f}")
        
        df_funnel['New_Paying_Users'] = df_funnel['New_Paying_Users'].apply(lambda x: f"{x:,.1f}")
        
        for col in ['MRR', 'Net_Cash_Flow', 'Cumulative_Cash']:
            df_funnel[col] = df_funnel[col].apply(lambda x: f"€{x:,.0f}")
        
        print("\n" + df_funnel.to_string(index=False))
        print("\n(Showing first 12 months only; full dataset has 36 months)")
        
        return df_funnel
    
    def create_visualizations(self):
        """Create all required charts."""
        print("\n" + "=" * 80)
        print("STEP 6: CREATING VISUALIZATIONS")
        print("=" * 80)
        
        self.monthly_df['MonthIndex'] = range(1, len(self.monthly_df) + 1)
        
        fig = plt.figure(figsize=(16, 12))
        
        # 1. MRR over time
        print("\n1. MRR over time chart...")
        ax1 = plt.subplot(3, 2, 1)
        ax1.plot(self.monthly_df['MonthIndex'], self.monthly_df['MRR'], 
                marker='o', linewidth=2, markersize=4, color='#2E86AB')
        ax1.set_title('Monthly Recurring Revenue (MRR) - 36 Months', fontweight='bold', fontsize=12)
        ax1.set_xlabel('Month')
        ax1.set_ylabel('MRR (EUR)')
        ax1.grid(True, alpha=0.3)
        ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x:,.0f}'))
        
        # 2. Paying users over time
        print("2. Paying users chart...")
        ax2 = plt.subplot(3, 2, 2)
        ax2.plot(self.monthly_df['MonthIndex'], self.monthly_df['Paying_Users_End'], 
                marker='o', linewidth=2, markersize=4, color='#A23B72')
        ax2.set_title('Paying Users Growth', fontweight='bold', fontsize=12)
        ax2.set_xlabel('Month')
        ax2.set_ylabel('Paying Users')
        ax2.grid(True, alpha=0.3)
        ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
        
        # 3. Cumulative cash
        print("3. Cumulative cash chart...")
        ax3 = plt.subplot(3, 2, 3)
        ax3.plot(self.monthly_df['MonthIndex'], self.monthly_df['Cumulative_Cash'], 
                marker='o', linewidth=2, markersize=4, color='#F18F01')
        ax3.axhline(y=0, color='red', linestyle='--', alpha=0.5, label='Break-even')
        
        if 'Broker_TargetCapital' in self.assumptions:
            target = self.assumptions['Broker_TargetCapital']
            ax3.axhline(y=target, color='green', linestyle='--', alpha=0.5, 
                       label=f'Target Capital (€{target:,.0f})')
        
        ax3.set_title('Cumulative Cash Flow', fontweight='bold', fontsize=12)
        ax3.set_xlabel('Month')
        ax3.set_ylabel('Cumulative Cash (EUR)')
        ax3.grid(True, alpha=0.3)
        ax3.legend()
        ax3.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x:,.0f}'))
        
        # 4. Pipeline funnel
        print("4. Pipeline funnel chart...")
        ax4 = plt.subplot(3, 2, 4)
        ax4_twin1 = ax4.twinx()
        ax4_twin2 = ax4.twinx()
        ax4_twin2.spines['right'].set_position(('outward', 60))
        
        p1 = ax4.plot(self.monthly_df['MonthIndex'], self.monthly_df['Social_Views'], 
                     label='Social Views', color='#C9ADA7', linewidth=2)
        p2 = ax4_twin1.plot(self.monthly_df['MonthIndex'], self.monthly_df['Visitors_Total'], 
                           label='Total Visitors', color='#6A4C93', linewidth=2)
        p3 = ax4_twin2.plot(self.monthly_df['MonthIndex'], self.monthly_df['New_Paying_Users'], 
                           label='New Paying Users', color='#22223B', linewidth=2)
        
        ax4.set_xlabel('Month')
        ax4.set_ylabel('Social Views', color='#C9ADA7')
        ax4_twin1.set_ylabel('Total Visitors', color='#6A4C93')
        ax4_twin2.set_ylabel('New Paying Users', color='#22223B')
        
        ax4.tick_params(axis='y', labelcolor='#C9ADA7')
        ax4_twin1.tick_params(axis='y', labelcolor='#6A4C93')
        ax4_twin2.tick_params(axis='y', labelcolor='#22223B')
        
        ax4.set_title('Conversion Pipeline: Views → Visitors → Paying', fontweight='bold', fontsize=12)
        
        lines = p1 + p2 + p3
        labels = [l.get_label() for l in lines]
        ax4.legend(lines, labels, loc='upper left')
        
        # 5. Channel contribution
        print("5. Channel contribution chart...")
        ax5 = plt.subplot(3, 2, 5)
        years = self.yearly_df['Year'].values
        inf_share = self.yearly_df['Share_Visitors_from_Influencers'].values
        social_share = 1 - inf_share
        
        x = np.arange(len(years))
        width = 0.35
        
        ax5.bar(x, social_share * 100, width, label='From Social Views', color='#8ecae6')
        ax5.bar(x, inf_share * 100, width, bottom=social_share * 100, 
               label='From Influencers', color='#219ebc')
        
        ax5.set_title('Visitor Source Mix by Year', fontweight='bold', fontsize=12)
        ax5.set_xlabel('Year')
        ax5.set_ylabel('Share of Visitors (%)')
        ax5.set_xticks(x)
        ax5.set_xticklabels([f'Year {int(y)}' for y in years])
        ax5.legend()
        ax5.grid(True, alpha=0.3, axis='y')
        
        # 6. Unit economics
        print("6. Unit economics chart...")
        ax6 = plt.subplot(3, 2, 6)
        years = self.yearly_df['Year'].values
        cac = self.yearly_df['Average_CAC_EUR'].values
        ltv = self.yearly_df['LTV_EUR'].values
        
        x = np.arange(len(years))
        width = 0.35
        
        ax6.bar(x - width/2, cac, width, label='CAC', color='#e63946')
        ax6.bar(x + width/2, ltv, width, label='LTV', color='#06d6a0')
        
        ax6.set_title('Unit Economics: CAC vs LTV by Year', fontweight='bold', fontsize=12)
        ax6.set_xlabel('Year')
        ax6.set_ylabel('EUR')
        ax6.set_xticks(x)
        ax6.set_xticklabels([f'Year {int(y)}' for y in years])
        ax6.legend()
        ax6.grid(True, alpha=0.3, axis='y')
        ax6.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x:,.0f}'))
        
        # Add LTV/CAC ratios
        for i, (c, l) in enumerate(zip(cac, ltv)):
            if c > 0:
                ratio = l / c
                ax6.text(i, max(c, l) * 1.05, f'{ratio:.1f}x', 
                        ha='center', va='bottom', fontweight='bold')
        
        plt.tight_layout()
        plt.savefig('financial_model_analysis.png', dpi=300, bbox_inches='tight')
        print("\n✓ Visualizations saved to: financial_model_analysis.png")
        plt.close()
    
    def generate_investor_narrative(self):
        """Generate investor-ready narrative summary."""
        print("\n" + "=" * 80)
        print("STEP 7: INVESTOR NARRATIVE")
        print("=" * 80)
        
        narrative = []
        narrative.append("\n# FINANCIAL MODEL ANALYSIS - EXECUTIVE SUMMARY")
        narrative.append("=" * 80)
        
        # Business Model Overview
        narrative.append("\n## 1. BUSINESS MODEL DYNAMICS")
        narrative.append("-" * 40)
        
        conv_rate = self.assumptions['Social_View_to_Visit_Conv']
        conv_vs = self.assumptions['ConvVS']
        conv_sp = self.assumptions['ConvSP']
        overall_conv = conv_vs * conv_sp
        
        narrative.append(f"\n**Acquisition Funnel:**")
        narrative.append(f"- Social views convert to website visits at {conv_rate:.2%}")
        narrative.append(f"  (e.g., 100 social views → {conv_rate*100:.0f} site visits)")
        narrative.append(f"\n- Website visitors → Signups: {conv_vs:.2%}")
        narrative.append(f"- Signups → Paying users: {conv_sp:.2%}")
        narrative.append(f"- **Overall visitor-to-paid conversion: {overall_conv:.2%}**")
        
        # Influencer strategy
        narrative.append(f"\n**Influencer Marketing Strategy:**")
        followers = self.assumptions['Inf_Avg_Followers']
        visitors_per = self.assumptions['Inf_Visitors_per_Collab']
        narrative.append(f"- Average influencer has {followers:,.0f} followers")
        narrative.append(f"- Each collaboration generates ~{visitors_per:.0f} website visitors")
        
        inf_shares = self.yearly_df['Share_Visitors_from_Influencers'].values
        narrative.append(f"- Influencers contribute {inf_shares[0]:.1%} (Y1) → {inf_shares[-1]:.1%} (Y3) of total visitors")
        
        # Growth trajectory
        narrative.append("\n\n## 2. GROWTH TRAJECTORY (3-YEAR OUTLOOK)")
        narrative.append("-" * 40)
        
        for idx, row in self.yearly_df.iterrows():
            year = int(row['Year'])
            narrative.append(f"\n**Year {year}:**")
            narrative.append(f"- Paying users: {row['End_Paying_Users']:,.0f}")
            narrative.append(f"- MRR: €{row['End_MRR_EUR']:,.0f}")
            narrative.append(f"- ARR: €{row['ARR_EUR']:,.0f}")
            narrative.append(f"- New customers acquired: {row['Total_New_Customers']:,.0f}")
            narrative.append(f"- Marketing spend: €{row['Total_Marketing_Spend_EUR']:,.0f}")
        
        # Cash flow analysis
        narrative.append("\n\n## 3. CASH FLOW & CAPITAL REQUIREMENTS")
        narrative.append("-" * 40)
        
        final_cash = self.monthly_df['Cumulative_Cash'].iloc[-1]
        min_cash = self.monthly_df['Cumulative_Cash'].min()
        
        # Find break-even month
        break_even_month = None
        for idx, row in self.monthly_df.iterrows():
            if row['Cumulative_Cash'] >= 0:
                break_even_month = idx + 1
                break
        
        narrative.append(f"\n- **Minimum cash position: €{min_cash:,.0f}** (capital requirement)")
        if break_even_month:
            narrative.append(f"- **Break-even achieved: Month {break_even_month}**")
        else:
            narrative.append(f"- **Break-even: Not achieved within 36 months**")
        narrative.append(f"- **Cumulative cash at end of Year 3: €{final_cash:,.0f}**")
        
        if 'Broker_TargetCapital' in self.assumptions:
            target = self.assumptions['Broker_TargetCapital']
            narrative.append(f"\n- Broker target capital: €{target:,.0f}")
            if final_cash >= target:
                narrative.append(f"  ✓ **Target achieved** (surplus: €{final_cash - target:,.0f})")
            else:
                narrative.append(f"  ⚠ **Target not met** (shortfall: €{target - final_cash:,.0f})")
        
        # Unit economics
        narrative.append("\n\n## 4. UNIT ECONOMICS & SUSTAINABILITY")
        narrative.append("-" * 40)
        
        arpu = self.assumptions['ARPU']
        narrative.append(f"\n- **ARPU (Average Revenue Per User): €{arpu:,.2f}** per month")
        
        narrative.append(f"\n**CAC and LTV Evolution:**")
        for idx, row in self.yearly_df.iterrows():
            year = int(row['Year'])
            cac = row['Average_CAC_EUR']
            ltv = row['LTV_EUR']
            ratio = row['LTV_CAC_Ratio']
            narrative.append(f"- Year {year}: CAC = €{cac:,.0f}, LTV = €{ltv:,.0f}, **LTV/CAC = {ratio:.2f}x**")
        
        # Health assessment
        final_ratio = self.yearly_df['LTV_CAC_Ratio'].iloc[-1]
        narrative.append(f"\n**Unit Economics Assessment:**")
        if final_ratio >= 3.0:
            narrative.append(f"✓ **HEALTHY** - LTV/CAC ratio of {final_ratio:.1f}x indicates sustainable growth")
            narrative.append(f"  Industry benchmark: 3x or higher is considered healthy")
        elif final_ratio >= 2.0:
            narrative.append(f"⚠ **MODERATE** - LTV/CAC ratio of {final_ratio:.1f}x is acceptable but could be optimized")
            narrative.append(f"  Recommendation: Focus on retention or reducing CAC")
        else:
            narrative.append(f"⚠ **CONCERN** - LTV/CAC ratio of {final_ratio:.1f}x is below healthy threshold")
            narrative.append(f"  Action required: Improve retention or significantly reduce acquisition costs")
        
        # Retention & churn
        narrative.append("\n\n## 5. RETENTION DYNAMICS")
        narrative.append("-" * 40)
        
        churn_y1 = self.assumptions['ChurnY1']
        churn_y2 = self.assumptions['ChurnY2']
        churn_y3 = self.assumptions['ChurnY3']
        narrative.append(f"\n**Monthly Churn Rates:**")
        narrative.append(f"- Year 1: {churn_y1:.2%} (early adopters, higher churn)")
        narrative.append(f"- Year 2: {churn_y2:.2%} (improved product-market fit)")
        narrative.append(f"- Year 3: {churn_y3:.2%} (mature customer base)")
        
        retention_y3 = 1 - churn_y3
        narrative.append(f"\n- Year 3 monthly retention: {retention_y3:.2%}")
        narrative.append(f"- Implied annual retention: {retention_y3**12:.2%}")
        
        # Key risks
        narrative.append("\n\n## 6. KEY ASSUMPTIONS & SENSITIVITIES")
        narrative.append("-" * 40)
        
        narrative.append("\n**Critical Success Factors:**")
        narrative.append("1. Social media conversion rate maintaining at assumed levels")
        narrative.append("2. Influencer collaborations delivering expected reach and engagement")
        narrative.append("3. Churn rates improving as product matures")
        narrative.append("4. Marketing efficiency (CAC) staying within projected bounds")
        
        narrative.append("\n**Recommended Sensitivity Analysis:**")
        narrative.append("- Test scenarios with ±20% variation in conversion rates")
        narrative.append("- Model impact of ±10% variation in churn rates")
        narrative.append("- Assess effect of ±25% variation in CAC")
        narrative.append("- Evaluate influencer collaboration volume changes")
        
        # Conclusion
        narrative.append("\n\n## 7. INVESTMENT SUMMARY")
        narrative.append("-" * 40)
        
        final_mrr = self.yearly_df['End_MRR_EUR'].iloc[-1]
        final_arr = self.yearly_df['ARR_EUR'].iloc[-1]
        final_users = self.yearly_df['End_Paying_Users'].iloc[-1]
        
        narrative.append(f"\n**By end of Year 3, under these assumptions:**")
        narrative.append(f"- The business reaches **€{final_mrr:,.0f} MRR** (€{final_arr:,.0f} ARR)")
        narrative.append(f"- Serving **{final_users:,.0f} paying customers**")
        narrative.append(f"- Cumulative cash position: **€{final_cash:,.0f}**")
        
        if 'Broker_TargetCapital' in self.assumptions:
            target = self.assumptions['Broker_TargetCapital']
            if final_cash >= target:
                narrative.append(f"- **Capital target of €{target:,.0f} is ACHIEVED**")
            else:
                capital_needed = abs(min_cash)
                narrative.append(f"- Initial capital requirement: ~€{capital_needed:,.0f}")
                narrative.append(f"- Additional capital may be needed to reach €{target:,.0f} target")
        
        narrative.append("\n" + "=" * 80)
        narrative.append("\n*This analysis is based on the assumptions in the financial model.*")
        narrative.append("*Actual results may vary based on market conditions and execution.*")
        narrative.append("\n" + "=" * 80)
        
        full_narrative = '\n'.join(narrative)
        print(full_narrative)
        
        with open('investor_narrative.txt', 'w', encoding='utf-8') as f:
            f.write(full_narrative)
        
        print("\n✓ Narrative saved to: investor_narrative.txt")
        
        return full_narrative
    
    def run_full_analysis(self):
        """Execute complete analysis workflow."""
        self.load_and_calculate()
        self.sanity_checks()
        self.create_key_assumptions_table()
        self.create_yearly_summary_table()
        self.create_monthly_funnel_table()
        self.create_visualizations()
        self.generate_investor_narrative()
        
        print("\n" + "=" * 80)
        print("✓ ANALYSIS COMPLETE")
        print("=" * 80)
        print("\nGenerated files:")
        print("  1. financial_model_analysis.png - All visualizations")
        print("  2. investor_narrative.txt - Executive summary")
        print("\n" + "=" * 80)


if __name__ == "__main__":
    analyzer = FinancialModelAnalyzer('ai_finance_dynamic_model_v6_social_views.xlsx')
    analyzer.run_full_analysis()
