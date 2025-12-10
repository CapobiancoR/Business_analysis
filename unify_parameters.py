"""
Script per unificare i parametri ChurnY1/Y2/Y3 e Other_Marketing_Budget_Y1/Y2/Y3
in parametri singoli: Churn_Rate e Other_Marketing_Budget
"""
import pandas as pd
from openpyxl import load_workbook

excel_path = 'ai_finance_dynamic_model_v7_channels.xlsx'

# Carica il workbook
wb = load_workbook(excel_path)
ws = wb['Model']

# Cerca e modifica i parametri
changes_made = []

for row in range(4, 100):
    param = ws.cell(row=row, column=2).value  # Column B = Parameter
    
    if param is None:
        continue
    
    # Unifica ChurnY1 → Churn_Rate (e rimuovi Y2, Y3)
    if param == 'ChurnY1':
        ws.cell(row=row, column=2, value='Churn_Rate')
        ws.cell(row=row, column=5, value='Monthly churn rate (unified for all years)')
        changes_made.append(f"Row {row}: ChurnY1 → Churn_Rate")
    elif param in ['ChurnY2', 'ChurnY3']:
        # Svuota queste righe (verranno ignorate)
        ws.cell(row=row, column=1, value='')
        ws.cell(row=row, column=2, value=f'[DEPRECATED_{param}]')
        ws.cell(row=row, column=5, value='DEPRECATED - use Churn_Rate instead')
        changes_made.append(f"Row {row}: {param} → DEPRECATED")
    
    # Unifica Other_Marketing_Budget_Y1 → Other_Marketing_Budget (e rimuovi Y2, Y3)
    if param == 'Other_Marketing_Budget_Y1':
        ws.cell(row=row, column=2, value='Other_Marketing_Budget')
        ws.cell(row=row, column=5, value='Other marketing budget per month (unified for all years)')
        changes_made.append(f"Row {row}: Other_Marketing_Budget_Y1 → Other_Marketing_Budget")
    elif param in ['Other_Marketing_Budget_Y2', 'Other_Marketing_Budget_Y3']:
        ws.cell(row=row, column=1, value='')
        ws.cell(row=row, column=2, value=f'[DEPRECATED_{param}]')
        ws.cell(row=row, column=5, value='DEPRECATED - use Other_Marketing_Budget instead')
        changes_made.append(f"Row {row}: {param} → DEPRECATED")

# Salva
wb.save(excel_path)

print("✓ Modifiche effettuate:")
for change in changes_made:
    print(f"  {change}")
print(f"\n✓ Salvato {excel_path}")
