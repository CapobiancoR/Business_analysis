"""
Script per aggiungere i nuovi parametri richiesti dai 4 fix all'Excel v7.

NUOVI PARAMETRI DA AGGIUNGERE:
1. Inf_Avg_Followers - media followers degli influencer
2. Inf_Reach_Rate - tasso reach influencer
3. Inf_Click_Rate - tasso click influencer
4. FollowerAds_CTR_to_Site - CTR delle follower ads verso il sito
5. Follower_Threshold_For_Click_Ads - già esiste (verificare)

PARAMETRI DA MANTENERE MA NON MODIFICABILI:
- Inf_Visitors_per_Collab (sarà calcolato, non editabile)
"""

import openpyxl
from openpyxl.styles import Alignment
import sys

def add_parameters_to_excel(filepath: str):
    """Aggiungi i nuovi parametri all'Excel v7."""
    
    try:
        wb = openpyxl.load_workbook(filepath)
        # Il file ha un solo sheet 'Model' con tutto dentro
        ws = wb.active
        
        print(f"📂 Caricato: {filepath}")
        print(f"   Sheet: {ws.title}")
        print(f"   Available sheets: {wb.sheetnames}")
        
        # Trova l'ultima riga con dati
        last_row = ws.max_row
        print(f"   Ultima riga: {last_row}")
        
        # Parametri da aggiungere
        new_params = [
            {
                'Category': 'Influencer',
                'Parameter': 'Inf_Avg_Followers',
                'Value': 50000,
                'Unit': 'followers',
                'Notes': 'Average follower count of influencers (for calculation)'
            },
            {
                'Category': 'Influencer',
                'Parameter': 'Inf_Reach_Rate',
                'Value': 0.3,
                'Unit': 'decimal',
                'Notes': 'Reach rate of influencer posts (30%)'
            },
            {
                'Category': 'Influencer',
                'Parameter': 'Inf_Click_Rate',
                'Value': 0.02,
                'Unit': 'decimal',
                'Notes': 'Click-through rate from influencer posts to site (2%)'
            },
            {
                'Category': 'Ads',
                'Parameter': 'FollowerAds_CTR_to_Site',
                'Value': 0.01,
                'Unit': 'decimal',
                'Notes': 'CTR from Follower Ads campaigns to website (1%)'
            },
        ]
        
        # Verifica se i parametri esistono già
        existing_params = set()
        for row in range(2, last_row + 1):
            param_name = ws.cell(row=row, column=2).value
            if param_name:
                existing_params.add(param_name)
        
        print(f"\n📊 Parametri esistenti: {len(existing_params)}")
        
        # Aggiungi i nuovi parametri
        current_row = last_row + 1
        added_count = 0
        
        for param in new_params:
            if param['Parameter'] in existing_params:
                print(f"⚠️  {param['Parameter']} - GIÀ ESISTENTE, salto")
                continue
            
            # Aggiungi riga
            ws.cell(row=current_row, column=1, value=param['Category'])
            ws.cell(row=current_row, column=2, value=param['Parameter'])
            ws.cell(row=current_row, column=3, value=param['Value'])
            ws.cell(row=current_row, column=4, value=param['Unit'])
            ws.cell(row=current_row, column=5, value=param['Notes'])
            
            # Allineamento
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).alignment = Alignment(
                    horizontal='left',
                    vertical='center'
                )
            
            print(f"✓ Riga {current_row}: {param['Parameter']} = {param['Value']}")
            current_row += 1
            added_count += 1
        
        # Verifica Follower_Threshold_For_Click_Ads
        print("\n🔍 Verifico Follower_Threshold_For_Click_Ads...")
        if 'Follower_Threshold_For_Click_Ads' in existing_params:
            print("✓ Follower_Threshold_For_Click_Ads - GIÀ ESISTENTE")
        else:
            # Aggiungi anche questo
            ws.cell(row=current_row, column=1, value='Ads')
            ws.cell(row=current_row, column=2, value='Follower_Threshold_For_Click_Ads')
            ws.cell(row=current_row, column=3, value=20000)
            ws.cell(row=current_row, column=4, value='followers')
            ws.cell(row=current_row, column=5, value='Threshold to switch from Follower Ads to Click Ads')
            
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).alignment = Alignment(
                    horizontal='left',
                    vertical='center'
                )
            
            print(f"✓ Riga {current_row}: Follower_Threshold_For_Click_Ads = 20000")
            added_count += 1
        
        # Salva il file aggiornato
        if added_count > 0:
            output_path = filepath.replace('.xlsx', '_UPDATED.xlsx')
            wb.save(output_path)
            print(f"\n✅ Excel aggiornato con successo!")
            print(f"   File salvato come: {output_path}")
            print(f"   Parametri aggiunti: {added_count}")
            print(f"\n⚠️  RINOMINA IL FILE DA _UPDATED.xlsx AL NOME ORIGINALE")
        else:
            print(f"\n✅ Nessun parametro da aggiungere, Excel già aggiornato")
        
        return True
        
    except Exception as e:
        print(f"\n❌ ERRORE: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    excel_file = 'ai_finance_dynamic_model_v7_channels.xlsx'
    
    print("=" * 80)
    print("AGGIUNTA NUOVI PARAMETRI PER FIX 1-4")
    print("=" * 80)
    print()
    
    success = add_parameters_to_excel(excel_file)
    
    sys.exit(0 if success else 1)
