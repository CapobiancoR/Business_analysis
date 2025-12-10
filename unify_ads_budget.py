"""
Script per unificare FollowerAds_Budget_Y1/Y2/Y3 in PaidAds_Monthly_Budget
"""
from openpyxl import load_workbook

excel_path = 'ai_finance_dynamic_model_v7_channels.xlsx'

# Carica il workbook
wb = load_workbook(excel_path)
ws = wb['Model']

# Cerca e modifica i parametri
changes_made = []

for row in range(4, 100):
    param = ws.cell(row=row, column=2).value  # Column B = Parameter
    
    if param is None:
        continue
    
    # Unifica FollowerAds_Budget_Y1 → PaidAds_Monthly_Budget (e rimuovi Y2, Y3)
    if param == 'FollowerAds_Budget_Y1':
        ws.cell(row=row, column=2, value='PaidAds_Monthly_Budget')
        ws.cell(row=row, column=5, value='Monthly budget for paid ads - both Follower and Click Ads (unified for all years)')
        changes_made.append(f"Row {row}: FollowerAds_Budget_Y1 → PaidAds_Monthly_Budget")
    elif param in ['FollowerAds_Budget_Y2', 'FollowerAds_Budget_Y3']:
        ws.cell(row=row, column=1, value='')
        ws.cell(row=row, column=2, value=f'[DEPRECATED_{param}]')
        ws.cell(row=row, column=5, value='DEPRECATED - use PaidAds_Monthly_Budget instead')
        changes_made.append(f"Row {row}: {param} → DEPRECATED")

# Salva
wb.save(excel_path)

print("✓ Modifiche effettuate:")
for change in changes_made:
    print(f"  {change}")
print(f"\n✓ Salvato {excel_path}")
