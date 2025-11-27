"""Inspect rows 50-94 to find monthly model"""
import pandas as pd
from openpyxl import load_workbook

filepath = 'ai_finance_dynamic_model_v6_social_views.xlsx'
wb = load_workbook(filepath, data_only=True)
sheet = wb["Model"]

print("=" * 80)
print("ROWS 50-94 (Monthly Model Area)")
print("=" * 80)

for row_idx in range(50, min(95, sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(15, sheet.max_column + 1)):
        cell = sheet.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is not None:
            val_str = str(val)[:20]  # Truncate long values
            row_data.append(val_str)
        else:
            row_data.append("")
    
    # Only print non-empty rows
    if any(row_data):
        print(f"Row {row_idx:3d}: {' | '.join(row_data)}")
