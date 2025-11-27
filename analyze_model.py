"""
Senior Data Analyst & Python Engineer: Financial Model Analysis
================================================================

This script analyzes the AI Finance Dynamic Model Excel file.
It extracts assumptions, monthly projections, yearly summaries,
performs sanity checks, and creates investor-ready visualizations.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from pathlib import Path
import warnings
import xlwings as xw
warnings.filterwarnings('ignore')

# Set visualization style
sns.set_style("whitegrid")
plt.rcParams['figure.figsize'] = (12, 6)
plt.rcParams['font.size'] = 10

# Flag to use xlwings for formula evaluation (Windows only)
USE_XLWINGS = False  # Set to True if xlwings is available and Excel is installed

class FinancialModelAnalyzer:
    """Analyzes Excel financial model with assumptions, monthly projections, and yearly summaries."""
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = None
        self.sheet = None
        self.assumptions = {}
        self.monthly_df = None
        self.yearly_df = None
        
    def load_excel(self):
        """Load Excel file and identify the Model sheet."""
        print("=" * 80)
        print("STEP 1: LOADING AND PARSING EXCEL FILE")
        print("=" * 80)
        
        # Load with openpyxl to preserve formulas
        self.workbook = load_workbook(self.filepath, data_only=True)
        
        # Find the Model sheet
        if "Model" in self.workbook.sheetnames:
            self.sheet = self.workbook["Model"]
            print(f"✓ Found 'Model' sheet")
        else:
            # Try first sheet
            self.sheet = self.workbook[self.workbook.sheetnames[0]]
            print(f"⚠ Using first sheet: {self.workbook.sheetnames[0]}")
        
        print(f"✓ Workbook loaded successfully")
        print(f"  Sheet name: {self.sheet.title}")
        print(f"  Max row: {self.sheet.max_row}, Max column: {self.sheet.max_column}")
        
    def extract_assumptions(self):
        """Extract assumptions from the top section of the sheet."""
        print("\n" + "=" * 80)
        print("STEP 2: EXTRACTING ASSUMPTIONS")
        print("=" * 80)
        
        # Read the entire sheet as a DataFrame to analyze structure
        df_full = pd.read_excel(self.filepath, sheet_name="Model" if "Model" in pd.ExcelFile(self.filepath).sheet_names else 0, header=None)
        
        # Find assumptions table (look for "Category", "Parameter", "Value" headers)
        assumptions_start = None
        for idx, row in df_full.iterrows():
            row_str = ' '.join([str(x).lower() for x in row.values if pd.notna(x)])
            if 'category' in row_str and 'parameter' in row_str and 'value' in row_str:
                assumptions_start = idx
                print(f"✓ Found assumptions header at row {idx}")
                break
        
        if assumptions_start is not None:
            # Read assumptions table
            assumptions_df = pd.read_excel(self.filepath, sheet_name="Model" if "Model" in pd.ExcelFile(self.filepath).sheet_names else 0, 
                                          header=assumptions_start)
            
            # Extract key-value pairs (Parameter -> Value)
            for idx, row in assumptions_df.iterrows():
                if pd.notna(row.get('Parameter')) and pd.notna(row.get('Value')):
                    param = str(row['Parameter']).strip()
                    value = row['Value']
                    if param and param != 'Parameter':  # Skip header repeats
                        self.assumptions[param] = value
            
            print(f"✓ Extracted {len(self.assumptions)} assumptions")
            
        return self.assumptions
    
    def extract_monthly_model(self):
        """Extract 36-month projections."""
        print("\n" + "=" * 80)
        print("STEP 2B: EXTRACTING MONTHLY MODEL (36 months)")
        print("=" * 80)
        
        # Read the entire sheet
        df_full = pd.read_excel(self.filepath, sheet_name="Model" if "Model" in pd.ExcelFile(self.filepath).sheet_names else 0, header=None)
        
        # Find monthly model header (look for "Year", "Month", "Social_Views", etc.)
        monthly_start = None
        for idx, row in df_full.iterrows():
            row_str = ' '.join([str(x).lower() for x in row.values if pd.notna(x)])
            if 'month' in row_str and ('social_views' in row_str or 'visitors' in row_str or 'mrr' in row_str):
                monthly_start = idx
                print(f"✓ Found monthly model header at row {idx}")
                break
        
        if monthly_start is not None:
            # Read monthly data
            self.monthly_df = pd.read_excel(self.filepath, sheet_name="Model" if "Model" in pd.ExcelFile(self.filepath).sheet_names else 0,
                                           header=monthly_start)
            
            # Clean column names
            self.monthly_df.columns = self.monthly_df.columns.str.strip()
            
            # Filter to 36 rows (3 years)
            self.monthly_df = self.monthly_df[self.monthly_df['Year'].notna()].head(36)
            
            print(f"✓ Extracted {len(self.monthly_df)} months of data")
            print(f"✓ Columns: {', '.join(self.monthly_df.columns[:8])}...")
            
        return self.monthly_df
    
    def extract_yearly_summary(self):
        """Extract yearly summary table."""
        print("\n" + "=" * 80)
        print("STEP 2C: EXTRACTING YEARLY SUMMARY")
        print("=" * 80)
        
        # Read the entire sheet
        df_full = pd.read_excel(self.filepath, sheet_name="Model" if "Model" in pd.ExcelFile(self.filepath).sheet_names else 0, header=None)
        
        # Find yearly summary (look for headers like "End_Paying_Users", "ARR_EUR", etc.)
        yearly_start = None
        for idx, row in df_full.iterrows():
            row_str = ' '.join([str(x).lower() for x in row.values if pd.notna(x)])
            if 'arr_eur' in row_str or ('ltv' in row_str and 'cac' in row_str):
                yearly_start = idx
                print(f"✓ Found yearly summary header at row {idx}")
                break
        
        if yearly_start is not None:
            # Read yearly data
            self.yearly_df = pd.read_excel(self.filepath, sheet_name="Model" if "Model" in pd.ExcelFile(self.filepath).sheet_names else 0,
                                          header=yearly_start)
            
            # Clean column names
            self.yearly_df.columns = self.yearly_df.columns.str.strip()
            
            # Filter to 3 years
            self.yearly_df = self.yearly_df[self.yearly_df['Year'].notna()].head(3)
            
            print(f"✓ Extracted {len(self.yearly_df)} years of summary data")
            print(f"✓ Columns: {', '.join(self.yearly_df.columns[:8])}...")
        
        return self.yearly_df
    
    def sanity_checks(self):
        """Perform data validation checks."""
        print("\n" + "=" * 80)
        print("STEP 3: SANITY CHECKS")
        print("=" * 80)
        
        issues = []
        
        # Check 1: Channel shares sum to 1.0
        print("\n1. Checking channel shares sum to ~1.0...")
        for year in [1, 2, 3]:
            share_key = f'Share_Sum_Y{year}'
            if share_key in self.assumptions:
                share_sum = self.assumptions[share_key]
                if abs(share_sum - 1.0) > 0.01:
                    issues.append(f"Year {year} channel shares sum to {share_sum:.3f} (expected ~1.0)")
                    print(f"  ⚠ Year {year}: {share_sum:.3f}")
                else:
                    print(f"  ✓ Year {year}: {share_sum:.3f}")
        
        # Check 2: Months are sequential
        print("\n2. Checking months are sequential (1-12 per year)...")
        if self.monthly_df is not None:
            for year in [1, 2, 3]:
                year_data = self.monthly_df[self.monthly_df['Year'] == year]
                months = sorted(year_data['Month'].values)
                expected = list(range(1, 13))
                if list(months) != expected:
                    issues.append(f"Year {year} months are not 1-12: {months}")
                    print(f"  ⚠ Year {year}: {months}")
                else:
                    print(f"  ✓ Year {year}: Months 1-12 present")
        
        # Check 3: Cumulative cash consistency
        print("\n3. Checking cumulative cash consistency...")
        if self.monthly_df is not None and 'Net_Cash_Flow' in self.monthly_df.columns and 'Cumulative_Cash' in self.monthly_df.columns:
            calc_cumulative = self.monthly_df['Net_Cash_Flow'].cumsum()
            actual_cumulative = self.monthly_df['Cumulative_Cash']
            max_diff = abs(calc_cumulative - actual_cumulative).max()
            if max_diff > 1.0:  # Allow 1 EUR rounding error
                issues.append(f"Cumulative cash mismatch (max diff: {max_diff:.2f} EUR)")
                print(f"  ⚠ Max difference: {max_diff:.2f} EUR")
            else:
                print(f"  ✓ Cumulative cash is consistent (max diff: {max_diff:.2f} EUR)")
        
        # Check 4: Visitors total = Visitors_from_Social + Inf_Visitors
        print("\n4. Checking visitor totals...")
        if self.monthly_df is not None:
            required_cols = ['Visitors_Total', 'Visitors_from_Social', 'Inf_Visitors']
            if all(col in self.monthly_df.columns for col in required_cols):
                calc_total = self.monthly_df['Visitors_from_Social'] + self.monthly_df['Inf_Visitors']
                actual_total = self.monthly_df['Visitors_Total']
                max_diff = abs(calc_total - actual_total).max()
                if max_diff > 1.0:
                    issues.append(f"Visitor totals mismatch (max diff: {max_diff:.2f})")
                    print(f"  ⚠ Max difference: {max_diff:.2f}")
                else:
                    print(f"  ✓ Visitor totals are consistent")
        
        # Summary
        print("\n" + "-" * 80)
        if issues:
            print(f"⚠ Found {len(issues)} potential issues:")
            for issue in issues:
                print(f"  - {issue}")
        else:
            print("✓ All sanity checks passed!")
        
        return issues
    
    def create_key_assumptions_table(self):
        """Generate markdown table of key assumptions."""
        print("\n" + "=" * 80)
        print("STEP 4A: KEY ASSUMPTIONS TABLE")
        print("=" * 80)
        
        # Key parameters to display
        key_params = [
            ('ARPU', 'Revenue'),
            ('GrossMargin', 'Revenue'),
            ('ConvVS', 'Conversion'),
            ('ConvSP', 'Conversion'),
            ('Base_Visitor_to_Paid_Conv', 'Conversion'),
            ('Social_View_to_Visit_Conv', 'Conversion'),
            ('ChurnY1', 'Retention'),
            ('ChurnY2', 'Retention'),
            ('ChurnY3', 'Retention'),
            ('CAC_Y1', 'CAC'),
            ('CAC_Y2', 'CAC'),
            ('CAC_Y3', 'CAC'),
            ('Inf_Avg_Followers', 'Influencer'),
            ('Inf_Reach_Rate', 'Influencer'),
            ('Inf_Click_Rate', 'Influencer'),
            ('Inf_Visitors_per_Collab', 'Influencer'),
            ('BaseFixedCost', 'Costs'),
            ('DataSub_Fee', 'Costs'),
            ('DataSub_MRR_Threshold', 'Costs'),
            ('XAPI_Fee', 'Costs'),
            ('XAPI_MRR_Threshold', 'Costs'),
            ('Broker_TargetCapital', 'Capital'),
        ]
        
        table_data = []
        for param, category in key_params:
            if param in self.assumptions:
                value = self.assumptions[param]
                # Format value
                if isinstance(value, (int, float)):
                    if param.startswith('Conv') or param.endswith('Rate') or 'Margin' in param:
                        formatted = f"{value:.2%}" if value < 1 else f"{value:.4f}"
                    elif 'Churn' in param:
                        formatted = f"{value:.2%}"
                    else:
                        formatted = f"{value:,.2f}"
                else:
                    formatted = str(value)
                table_data.append([category, param, formatted])
        
        df_table = pd.DataFrame(table_data, columns=['Category', 'Parameter', 'Value'])
        
        print("\n" + df_table.to_string(index=False))
        
        return df_table
    
    def create_yearly_summary_table(self):
        """Generate clean yearly summary table."""
        print("\n" + "=" * 80)
        print("STEP 4B: YEARLY SUMMARY TABLE")
        print("=" * 80)
        
        if self.yearly_df is not None:
            # Select key columns
            key_cols = ['Year', 'End_Paying_Users', 'End_MRR_EUR', 'ARR_EUR', 
                       'Total_New_Customers', 'Total_Marketing_Spend_EUR', 'Average_CAC_EUR',
                       'LTV_EUR', 'LTV_CAC_Ratio', 'Cumulative_Cash_EndOfYear',
                       'Share_Visitors_from_Influencers', 'Total_Social_Views']
            
            # Filter available columns
            available_cols = [col for col in key_cols if col in self.yearly_df.columns]
            df_summary = self.yearly_df[available_cols].copy()
            
            # Format numeric columns
            for col in df_summary.columns:
                if col != 'Year' and df_summary[col].dtype in ['float64', 'int64']:
                    if 'Ratio' in col or 'Share' in col:
                        df_summary[col] = df_summary[col].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '')
                    elif 'EUR' in col or 'Marketing' in col or 'Cash' in col:
                        df_summary[col] = df_summary[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '')
                    else:
                        df_summary[col] = df_summary[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '')
            
            print("\n" + df_summary.to_string(index=False))
            
            return df_summary
        
        return None
    
    def create_monthly_funnel_table(self):
        """Generate compact monthly funnel table."""
        print("\n" + "=" * 80)
        print("STEP 4C: MONTHLY FUNNEL TABLE (First 12 months sample)")
        print("=" * 80)
        
        if self.monthly_df is not None:
            # Select key columns
            key_cols = ['Year', 'Month', 'Social_Views', 'Visitors_from_Social', 
                       'Inf_Visitors', 'Visitors_Total', 'New_Paying_Users', 
                       'MRR', 'Net_Cash_Flow', 'Cumulative_Cash']
            
            # Filter available columns
            available_cols = [col for col in key_cols if col in self.monthly_df.columns]
            df_funnel = self.monthly_df[available_cols].head(12).copy()
            
            # Format numeric columns
            for col in df_funnel.columns:
                if col not in ['Year', 'Month'] and df_funnel[col].dtype in ['float64', 'int64']:
                    if 'MRR' in col or 'Cash' in col:
                        df_funnel[col] = df_funnel[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '')
                    else:
                        df_funnel[col] = df_funnel[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '')
            
            print("\n" + df_funnel.to_string(index=False))
            print("\n(Showing first 12 months only; full dataset has 36 months)")
            
            return df_funnel
        
        return None
    
    def create_visualizations(self):
        """Create all required charts."""
        print("\n" + "=" * 80)
        print("STEP 5: CREATING VISUALIZATIONS")
        print("=" * 80)
        
        if self.monthly_df is None:
            print("⚠ Cannot create visualizations: monthly data not available")
            return
        
        # Create month index for x-axis
        self.monthly_df['MonthIndex'] = range(1, len(self.monthly_df) + 1)
        
        # Create figure with subplots
        fig = plt.figure(figsize=(16, 12))
        
        # 1. MRR over time
        print("\n1. Creating MRR over time chart...")
        ax1 = plt.subplot(3, 2, 1)
        if 'MRR' in self.monthly_df.columns:
            ax1.plot(self.monthly_df['MonthIndex'], self.monthly_df['MRR'], 
                    marker='o', linewidth=2, markersize=4, color='#2E86AB')
            ax1.set_title('Monthly Recurring Revenue (MRR) - 36 Months', fontweight='bold', fontsize=12)
            ax1.set_xlabel('Month')
            ax1.set_ylabel('MRR (EUR)')
            ax1.grid(True, alpha=0.3)
            ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x:,.0f}'))
        
        # 2. Paying users over time
        print("2. Creating paying users chart...")
        ax2 = plt.subplot(3, 2, 2)
        if 'Paying_Users_End' in self.monthly_df.columns:
            ax2.plot(self.monthly_df['MonthIndex'], self.monthly_df['Paying_Users_End'], 
                    marker='o', linewidth=2, markersize=4, color='#A23B72')
            ax2.set_title('Paying Users Growth', fontweight='bold', fontsize=12)
            ax2.set_xlabel('Month')
            ax2.set_ylabel('Paying Users')
            ax2.grid(True, alpha=0.3)
            ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
        
        # 3. Cumulative cash over time
        print("3. Creating cumulative cash chart...")
        ax3 = plt.subplot(3, 2, 3)
        if 'Cumulative_Cash' in self.monthly_df.columns:
            ax3.plot(self.monthly_df['MonthIndex'], self.monthly_df['Cumulative_Cash'], 
                    marker='o', linewidth=2, markersize=4, color='#F18F01')
            ax3.axhline(y=0, color='red', linestyle='--', alpha=0.5, label='Break-even')
            
            # Add Broker_TargetCapital line if available
            if 'Broker_TargetCapital' in self.assumptions:
                target = self.assumptions['Broker_TargetCapital']
                ax3.axhline(y=target, color='green', linestyle='--', alpha=0.5, 
                           label=f'Target Capital (€{target:,.0f})')
            
            ax3.set_title('Cumulative Cash Flow', fontweight='bold', fontsize=12)
            ax3.set_xlabel('Month')
            ax3.set_ylabel('Cumulative Cash (EUR)')
            ax3.grid(True, alpha=0.3)
            ax3.legend()
            ax3.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x:,.0f}'))
        
        # 4. Pipeline funnel: Social views → Visitors → Paying users
        print("4. Creating pipeline funnel chart...")
        ax4 = plt.subplot(3, 2, 4)
        if all(col in self.monthly_df.columns for col in ['Social_Views', 'Visitors_Total', 'New_Paying_Users']):
            ax4_twin1 = ax4.twinx()
            ax4_twin2 = ax4.twinx()
            
            # Offset the right spine
            ax4_twin2.spines['right'].set_position(('outward', 60))
            
            p1 = ax4.plot(self.monthly_df['MonthIndex'], self.monthly_df['Social_Views'], 
                         label='Social Views', color='#C9ADA7', linewidth=2)
            p2 = ax4_twin1.plot(self.monthly_df['MonthIndex'], self.monthly_df['Visitors_Total'], 
                               label='Total Visitors', color='#6A4C93', linewidth=2)
            p3 = ax4_twin2.plot(self.monthly_df['MonthIndex'], self.monthly_df['New_Paying_Users'], 
                               label='New Paying Users', color='#22223B', linewidth=2)
            
            ax4.set_xlabel('Month')
            ax4.set_ylabel('Social Views', color='#C9ADA7')
            ax4_twin1.set_ylabel('Total Visitors', color='#6A4C93')
            ax4_twin2.set_ylabel('New Paying Users', color='#22223B')
            
            ax4.tick_params(axis='y', labelcolor='#C9ADA7')
            ax4_twin1.tick_params(axis='y', labelcolor='#6A4C93')
            ax4_twin2.tick_params(axis='y', labelcolor='#22223B')
            
            ax4.set_title('Conversion Pipeline: Views → Visitors → Paying', fontweight='bold', fontsize=12)
            
            # Combine legends
            lines = p1 + p2 + p3
            labels = [l.get_label() for l in lines]
            ax4.legend(lines, labels, loc='upper left')
        
        # 5. Channel contribution (influencers vs social)
        print("5. Creating channel contribution chart...")
        ax5 = plt.subplot(3, 2, 5)
        if self.yearly_df is not None and 'Share_Visitors_from_Influencers' in self.yearly_df.columns:
            years = self.yearly_df['Year'].values
            inf_share = self.yearly_df['Share_Visitors_from_Influencers'].values
            social_share = 1 - inf_share
            
            x = np.arange(len(years))
            width = 0.35
            
            ax5.bar(x, social_share * 100, width, label='From Social Views', color='#8ecae6')
            ax5.bar(x, inf_share * 100, width, bottom=social_share * 100, 
                   label='From Influencers', color='#219ebc')
            
            ax5.set_title('Visitor Source Mix by Year', fontweight='bold', fontsize=12)
            ax5.set_xlabel('Year')
            ax5.set_ylabel('Share of Visitors (%)')
            ax5.set_xticks(x)
            ax5.set_xticklabels([f'Year {int(y)}' for y in years])
            ax5.legend()
            ax5.grid(True, alpha=0.3, axis='y')
        
        # 6. Unit economics (CAC vs LTV)
        print("6. Creating unit economics chart...")
        ax6 = plt.subplot(3, 2, 6)
        if self.yearly_df is not None:
            if all(col in self.yearly_df.columns for col in ['Year', 'Average_CAC_EUR', 'LTV_EUR']):
                years = self.yearly_df['Year'].values
                cac = self.yearly_df['Average_CAC_EUR'].values
                ltv = self.yearly_df['LTV_EUR'].values
                
                x = np.arange(len(years))
                width = 0.35
                
                ax6.bar(x - width/2, cac, width, label='CAC', color='#e63946')
                ax6.bar(x + width/2, ltv, width, label='LTV', color='#06d6a0')
                
                ax6.set_title('Unit Economics: CAC vs LTV by Year', fontweight='bold', fontsize=12)
                ax6.set_xlabel('Year')
                ax6.set_ylabel('EUR')
                ax6.set_xticks(x)
                ax6.set_xticklabels([f'Year {int(y)}' for y in years])
                ax6.legend()
                ax6.grid(True, alpha=0.3, axis='y')
                ax6.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x:,.0f}'))
                
                # Add LTV/CAC ratios as text
                for i, (c, l) in enumerate(zip(cac, ltv)):
                    if c > 0:
                        ratio = l / c
                        ax6.text(i, max(c, l) * 1.05, f'{ratio:.1f}x', 
                                ha='center', va='bottom', fontweight='bold')
        
        plt.tight_layout()
        plt.savefig('financial_model_analysis.png', dpi=300, bbox_inches='tight')
        print("\n✓ Visualizations saved to: financial_model_analysis.png")
        plt.close()
    
    def generate_investor_narrative(self):
        """Generate investor-ready narrative summary."""
        print("\n" + "=" * 80)
        print("STEP 6: INVESTOR NARRATIVE")
        print("=" * 80)
        
        narrative = []
        narrative.append("\n# FINANCIAL MODEL ANALYSIS - EXECUTIVE SUMMARY")
        narrative.append("=" * 80)
        
        # Business Model Overview
        narrative.append("\n## 1. BUSINESS MODEL DYNAMICS")
        narrative.append("-" * 40)
        
        if 'Social_View_to_Visit_Conv' in self.assumptions:
            conv_rate = self.assumptions['Social_View_to_Visit_Conv']
            narrative.append(f"\n**Acquisition Funnel:**")
            narrative.append(f"- Social views convert to website visits at {conv_rate:.2%}")
            narrative.append(f"  (e.g., 100 social views → {conv_rate*100:.0f} site visits)")
        
        if all(k in self.assumptions for k in ['ConvVS', 'ConvSP']):
            conv_vs = self.assumptions['ConvVS']
            conv_sp = self.assumptions['ConvSP']
            overall_conv = conv_vs * conv_sp
            narrative.append(f"\n- Website visitors → Signups: {conv_vs:.2%}")
            narrative.append(f"- Signups → Paying users: {conv_sp:.2%}")
            narrative.append(f"- **Overall visitor-to-paid conversion: {overall_conv:.2%}**")
        
        # Influencer strategy
        narrative.append(f"\n**Influencer Marketing Strategy:**")
        if 'Inf_Avg_Followers' in self.assumptions:
            followers = self.assumptions['Inf_Avg_Followers']
            narrative.append(f"- Average influencer has {followers:,.0f} followers")
        
        if 'Inf_Visitors_per_Collab' in self.assumptions:
            visitors_per = self.assumptions['Inf_Visitors_per_Collab']
            narrative.append(f"- Each collaboration generates ~{visitors_per:.0f} website visitors")
        
        if self.yearly_df is not None and 'Share_Visitors_from_Influencers' in self.yearly_df.columns:
            inf_shares = self.yearly_df['Share_Visitors_from_Influencers'].values
            narrative.append(f"- Influencers contribute {inf_shares[0]:.1%} (Y1) → {inf_shares[-1]:.1%} (Y3) of total visitors")
        
        # Growth trajectory
        narrative.append("\n\n## 2. GROWTH TRAJECTORY (3-YEAR OUTLOOK)")
        narrative.append("-" * 40)
        
        if self.yearly_df is not None:
            for idx, row in self.yearly_df.iterrows():
                year = int(row['Year'])
                narrative.append(f"\n**Year {year}:**")
                
                if 'End_Paying_Users' in row:
                    narrative.append(f"- Paying users: {row['End_Paying_Users']:,.0f}")
                
                if 'End_MRR_EUR' in row:
                    narrative.append(f"- MRR: €{row['End_MRR_EUR']:,.0f}")
                
                if 'ARR_EUR' in row:
                    narrative.append(f"- ARR: €{row['ARR_EUR']:,.0f}")
                
                if 'Total_New_Customers' in row:
                    narrative.append(f"- New customers acquired: {row['Total_New_Customers']:,.0f}")
                
                if 'Total_Marketing_Spend_EUR' in row:
                    narrative.append(f"- Marketing spend: €{row['Total_Marketing_Spend_EUR']:,.0f}")
        
        # Cash flow analysis
        narrative.append("\n\n## 3. CASH FLOW & CAPITAL REQUIREMENTS")
        narrative.append("-" * 40)
        
        if self.monthly_df is not None and 'Cumulative_Cash' in self.monthly_df.columns:
            final_cash = self.monthly_df['Cumulative_Cash'].iloc[-1]
            min_cash = self.monthly_df['Cumulative_Cash'].min()
            
            # Find break-even month
            break_even_month = None
            for idx, row in self.monthly_df.iterrows():
                if row['Cumulative_Cash'] >= 0:
                    break_even_month = idx + 1
                    break
            
            narrative.append(f"\n- **Minimum cash position: €{min_cash:,.0f}** (capital requirement)")
            if break_even_month:
                narrative.append(f"- **Break-even achieved: Month {break_even_month}**")
            else:
                narrative.append(f"- **Break-even: Not achieved within 36 months**")
            narrative.append(f"- **Cumulative cash at end of Year 3: €{final_cash:,.0f}**")
            
            if 'Broker_TargetCapital' in self.assumptions:
                target = self.assumptions['Broker_TargetCapital']
                narrative.append(f"\n- Broker target capital: €{target:,.0f}")
                if final_cash >= target:
                    narrative.append(f"  ✓ **Target achieved** (surplus: €{final_cash - target:,.0f})")
                else:
                    narrative.append(f"  ⚠ **Target not met** (shortfall: €{target - final_cash:,.0f})")
        
        # Unit economics
        narrative.append("\n\n## 4. UNIT ECONOMICS & SUSTAINABILITY")
        narrative.append("-" * 40)
        
        if 'ARPU' in self.assumptions:
            arpu = self.assumptions['ARPU']
            narrative.append(f"\n- **ARPU (Average Revenue Per User): €{arpu:,.2f}** per month")
        
        if self.yearly_df is not None and all(col in self.yearly_df.columns for col in ['Average_CAC_EUR', 'LTV_EUR', 'LTV_CAC_Ratio']):
            narrative.append(f"\n**CAC and LTV Evolution:**")
            for idx, row in self.yearly_df.iterrows():
                year = int(row['Year'])
                cac = row['Average_CAC_EUR']
                ltv = row['LTV_EUR']
                ratio = row['LTV_CAC_Ratio']
                narrative.append(f"- Year {year}: CAC = €{cac:,.0f}, LTV = €{ltv:,.0f}, **LTV/CAC = {ratio:.2f}x**")
            
            # Health assessment
            final_ratio = self.yearly_df['LTV_CAC_Ratio'].iloc[-1]
            narrative.append(f"\n**Unit Economics Assessment:**")
            if final_ratio >= 3.0:
                narrative.append(f"✓ **HEALTHY** - LTV/CAC ratio of {final_ratio:.1f}x indicates sustainable growth")
                narrative.append(f"  Industry benchmark: 3x or higher is considered healthy")
            elif final_ratio >= 2.0:
                narrative.append(f"⚠ **MODERATE** - LTV/CAC ratio of {final_ratio:.1f}x is acceptable but could be optimized")
                narrative.append(f"  Recommendation: Focus on retention or reducing CAC")
            else:
                narrative.append(f"⚠ **CONCERN** - LTV/CAC ratio of {final_ratio:.1f}x is below healthy threshold")
                narrative.append(f"  Action required: Improve retention or significantly reduce acquisition costs")
        
        # Retention & churn
        narrative.append("\n\n## 5. RETENTION DYNAMICS")
        narrative.append("-" * 40)
        
        if all(k in self.assumptions for k in ['ChurnY1', 'ChurnY2', 'ChurnY3']):
            churn_y1 = self.assumptions['ChurnY1']
            churn_y2 = self.assumptions['ChurnY2']
            churn_y3 = self.assumptions['ChurnY3']
            narrative.append(f"\n**Monthly Churn Rates:**")
            narrative.append(f"- Year 1: {churn_y1:.2%} (early adopters, higher churn)")
            narrative.append(f"- Year 2: {churn_y2:.2%} (improved product-market fit)")
            narrative.append(f"- Year 3: {churn_y3:.2%} (mature customer base)")
            
            # Calculate retention
            retention_y3 = 1 - churn_y3
            narrative.append(f"\n- Year 3 monthly retention: {retention_y3:.2%}")
            narrative.append(f"- Implied annual retention: {retention_y3**12:.2%}")
        
        # Key risks & opportunities
        narrative.append("\n\n## 6. KEY ASSUMPTIONS & SENSITIVITIES")
        narrative.append("-" * 40)
        
        narrative.append("\n**Critical Success Factors:**")
        narrative.append("1. Social media conversion rate maintaining at assumed levels")
        narrative.append("2. Influencer collaborations delivering expected reach and engagement")
        narrative.append("3. Churn rates improving as product matures")
        narrative.append("4. Marketing efficiency (CAC) staying within projected bounds")
        
        narrative.append("\n**Recommended Sensitivity Analysis:**")
        narrative.append("- Test scenarios with ±20% variation in conversion rates")
        narrative.append("- Model impact of ±10% variation in churn rates")
        narrative.append("- Assess effect of ±25% variation in CAC")
        narrative.append("- Evaluate influencer collaboration volume changes")
        
        # Conclusion
        narrative.append("\n\n## 7. INVESTMENT SUMMARY")
        narrative.append("-" * 40)
        
        if self.monthly_df is not None and self.yearly_df is not None:
            final_mrr = self.yearly_df['End_MRR_EUR'].iloc[-1]
            final_arr = self.yearly_df['ARR_EUR'].iloc[-1]
            final_users = self.yearly_df['End_Paying_Users'].iloc[-1]
            final_cash = self.monthly_df['Cumulative_Cash'].iloc[-1]
            
            narrative.append(f"\n**By end of Year 3, under these assumptions:**")
            narrative.append(f"- The business reaches **€{final_mrr:,.0f} MRR** (€{final_arr:,.0f} ARR)")
            narrative.append(f"- Serving **{final_users:,.0f} paying customers**")
            narrative.append(f"- Cumulative cash position: **€{final_cash:,.0f}**")
            
            if 'Broker_TargetCapital' in self.assumptions:
                target = self.assumptions['Broker_TargetCapital']
                if final_cash >= target:
                    narrative.append(f"- **Capital target of €{target:,.0f} is ACHIEVED**")
                else:
                    capital_needed = abs(min_cash)
                    narrative.append(f"- Initial capital requirement: ~€{capital_needed:,.0f}")
                    narrative.append(f"- Additional capital may be needed to reach €{target:,.0f} target")
        
        narrative.append("\n" + "=" * 80)
        narrative.append("\n*This analysis is based on the assumptions in the financial model.*")
        narrative.append("*Actual results may vary based on market conditions and execution.*")
        narrative.append("\n" + "=" * 80)
        
        full_narrative = '\n'.join(narrative)
        print(full_narrative)
        
        # Save to file
        with open('investor_narrative.txt', 'w', encoding='utf-8') as f:
            f.write(full_narrative)
        
        print("\n✓ Narrative saved to: investor_narrative.txt")
        
        return full_narrative
    
    def run_full_analysis(self):
        """Execute complete analysis workflow."""
        self.load_excel()
        self.extract_assumptions()
        self.extract_monthly_model()
        self.extract_yearly_summary()
        self.sanity_checks()
        self.create_key_assumptions_table()
        self.create_yearly_summary_table()
        self.create_monthly_funnel_table()
        self.create_visualizations()
        self.generate_investor_narrative()
        
        print("\n" + "=" * 80)
        print("✓ ANALYSIS COMPLETE")
        print("=" * 80)
        print("\nGenerated files:")
        print("  1. financial_model_analysis.png - All visualizations")
        print("  2. investor_narrative.txt - Executive summary")
        print("\n" + "=" * 80)


if __name__ == "__main__":
    # Run the analysis
    analyzer = FinancialModelAnalyzer('ai_finance_dynamic_model_v6_social_views.xlsx')
    analyzer.run_full_analysis()
