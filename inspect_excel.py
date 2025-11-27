"""Quick inspection script to understand Excel structure"""
import pandas as pd
from openpyxl import load_workbook

filepath = 'ai_finance_dynamic_model_v6_social_views.xlsx'

# Load with openpyxl
wb = load_workbook(filepath, data_only=True)
sheet = wb["Model"]

print("=" * 80)
print("EXCEL STRUCTURE INSPECTION")
print("=" * 80)

print(f"\nSheet: {sheet.title}")
print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")

# Print first 50 rows to understand structure
print("\n" + "=" * 80)
print("FIRST 50 ROWS (showing first 10 columns)")
print("=" * 80)

for row_idx in range(1, min(51, sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(11, sheet.max_column + 1)):
        cell = sheet.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is not None:
            val_str = str(val)[:30]  # Truncate long values
            row_data.append(val_str)
        else:
            row_data.append("")
    
    # Only print non-empty rows
    if any(row_data):
        print(f"Row {row_idx:3d}: {' | '.join(row_data)}")

# Also try reading with pandas to see what it detects
print("\n" + "=" * 80)
print("PANDAS AUTO-DETECTION TEST")
print("=" * 80)

df = pd.read_excel(filepath, sheet_name="Model", header=None, nrows=50)
print(df.to_string())
