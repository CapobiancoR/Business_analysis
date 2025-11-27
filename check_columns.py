"""Check all columns in monthly model"""
import pandas as pd
from openpyxl import load_workbook

filepath = 'ai_finance_dynamic_model_v6_social_views.xlsx'
wb = load_workbook(filepath, data_only=True)
sheet = wb["Model"]

print("=" * 80)
print("ROW 52 - ALL COLUMNS (Monthly Model Header)")
print("=" * 80)

for col_idx in range(1, sheet.max_column + 1):
    cell = sheet.cell(row=52, column=col_idx)
    if cell.value:
        print(f"Column {col_idx}: {cell.value}")

print("\n" + "=" * 80)
print("ROW 53 - FIRST DATA ROW (All columns)")
print("=" * 80)

for col_idx in range(1, sheet.max_column + 1):
    cell = sheet.cell(row=53, column=col_idx)
    if cell.value is not None:
        print(f"Column {col_idx}: {cell.value}")
