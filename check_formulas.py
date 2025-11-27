"""Check if cells contain formulas"""
from openpyxl import load_workbook

filepath = 'ai_finance_dynamic_model_v6_social_views.xlsx'

# Load WITHOUT data_only to see formulas
wb = load_workbook(filepath, data_only=False)
sheet = wb["Model"]

print("=" * 80)
print("CHECKING FOR FORMULAS IN ROW 53")
print("=" * 80)

for col_idx in range(1, 23):
    cell = sheet.cell(row=53, column=col_idx)
    header = sheet.cell(row=52, column=col_idx).value
    
    if cell.value is not None:
        # Check if it's a formula
        if isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"{header:25s} : FORMULA = {cell.value[:80]}")
        else:
            print(f"{header:25s} : VALUE = {cell.value}")

print("\n" + "=" * 80)
print("CHECKING DERIVED ASSUMPTIONS (formulas)")
print("=" * 80)

# Check some derived parameters
for row_idx in range(4, 50):
    param_cell = sheet.cell(row=row_idx, column=2)
    value_cell = sheet.cell(row=row_idx, column=3)
    
    if param_cell.value in ['Base_Visitor_to_Paid_Conv', 'Share_Sum_Y1', 'CAC_Y1', 'Inf_Visitors_per_Collab']:
        if isinstance(value_cell.value, str) and value_cell.value.startswith('='):
            print(f"{param_cell.value:30s} : {value_cell.value}")
        else:
            print(f"{param_cell.value:30s} : {value_cell.value} (static value)")
