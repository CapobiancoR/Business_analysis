"""
Script per aggiungere il parametro FixedCost_Annual_Growth all'Excel v7.
"""
import pandas as pd
from openpyxl import load_workbook

excel_path = 'ai_finance_dynamic_model_v7_channels.xlsx'

# Carica il workbook
wb = load_workbook(excel_path)
ws = wb['Model']

# Trova l'ultima riga delle assumptions (cerca BaseFixedCost e aggiungi dopo)
target_row = None
for row in range(4, 100):
    cell_value = ws.cell(row=row, column=2).value  # Column B = Parameter
    if cell_value == 'BaseFixedCost':
        target_row = row + 1
        break

if target_row is None:
    print("ERROR: Parametro BaseFixedCost non trovato!")
    exit(1)

# Sposta tutte le righe successive in giù di 1
ws.insert_rows(target_row)

# Aggiungi il nuovo parametro
ws.cell(row=target_row, column=1, value='Costs')  # Category
ws.cell(row=target_row, column=2, value='FixedCost_Annual_Growth')  # Parameter
ws.cell(row=target_row, column=3, value=0.05)  # Value (5%)
ws.cell(row=target_row, column=4, value='% annual')  # Unit
ws.cell(row=target_row, column=5, value='Annual growth rate of fixed costs (e.g., 0.05 = 5%)')  # Notes

print(f"✓ Aggiunto parametro 'FixedCost_Annual_Growth' alla riga {target_row}")

# Salva
wb.save(excel_path)
print(f"✓ Salvato {excel_path}")
